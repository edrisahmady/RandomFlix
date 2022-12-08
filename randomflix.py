import random
from openpyxl import load_workbook
import pandas as pd
import sys
import numpy as np


def main():
	
	
	try:
		filename=sys.argv[1] # makes sure error is displayed if no command line arguments are given
	except IndexError:
		print("ERROR: PLEASE run this program with a .csv file")
		sys.exit(1)

	if len(sys.argv) > 2: # gives error if more than one command line argument is given
		print("ERROR: Please run this program with just ONE file. That being a .csv file!")
		sys.exit(1)
	
	

	read_file = pd.read_csv(filename)  ## reads file name provided
	read_file.to_excel("watchlist.xlsx", index = None, header=True) #converst csv file to excel





	wb=load_workbook("watchlist.xlsx")      ## Loads Excel file
	sheet=wb.active         ## Points to the most active page, since there is only one active sheet there is no need to specify anything else
	x = random.randint(2,sheet.max_row)     ## Determains a random number to represent row of movie

	print("Hello! Today's Random movie is:")
	print("===================================")


	headings=[]
	for data in sheet["1"]:               ## prints out random row for movie
		headings.append(data.value)

	# print(headings)


	details = []
	for data in sheet[x]:                  ## prints out random row for movie
		details.append(data.value)

	# print(details)


	print(np.array(list(zip(headings,details))))            ## puts the two lists together in a grid format for better viewing














main()
